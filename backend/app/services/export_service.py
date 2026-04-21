from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date


class ExcelExporter:
    """Generate Excel reports from data."""

    @staticmethod
    def create_report(title: str, headers: list[str], rows: list[list], sheet_name: str = "Отчёт") -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Arial", size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Date
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        date_cell = ws.cell(row=2, column=1, value=f"Дата: {date.today().strftime('%d.%m.%Y')}")
        date_cell.font = Font(name="Arial", size=10, color="666666")
        date_cell.alignment = Alignment(horizontal="center")

        # Headers
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
            top=Side(style="thin", color="E0E0E0"),
            bottom=Side(style="thin", color="E0E0E0"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")

        # Auto-width
        for col in ws.columns:
            max_length = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
